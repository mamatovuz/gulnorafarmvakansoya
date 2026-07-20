"""Excel (.xlsx) hisobotlarni tayyorlash."""
from datetime import datetime
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from aiogram.types import BufferedInputFile

from database.db import STATUS_LABELS

_HEADER_FILL = PatternFill("solid", fgColor="2E7D32")
_HEADER_FONT = Font(bold=True, color="FFFFFF")


def _autosize(ws):
    for col in ws.columns:
        width = 10
        letter = get_column_letter(col[0].column)
        for cell in col:
            value = "" if cell.value is None else str(cell.value)
            width = max(width, min(len(value) + 2, 50))
        ws.column_dimensions[letter].width = width


def _write_sheet(ws, headers, rows):
    ws.append(headers)
    for i, cell in enumerate(ws[1], start=1):
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.freeze_panes = "A2"
    for row in rows:
        ws.append(row)
    _autosize(ws)


def _finish(wb, prefix):
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    from utils import now_tk
    stamp = now_tk().strftime("%Y%m%d_%H%M")
    return BufferedInputFile(buf.read(), filename=f"{prefix}_{stamp}.xlsx")


def build_applications_xlsx(apps):
    wb = Workbook()
    ws = wb.active
    ws.title = "Arizalar"
    headers = [
        "#", "Ism-sharif", "Lavozim", "Filial", "Status", "Telefon",
        "Shahar", "Tuman", "Tug'ilgan sana", "Ma'lumot", "Tajriba",
        "Forma", "Kutilayotgan maosh", "Sana",
    ]
    uniform = {"yes": "Bor", "no": "Yo'q"}
    rows = []
    for a in apps:
        rows.append([
            a.get("id"),
            a.get("full_name") or "-",
            a.get("vacancy_title") or a.get("position") or "-",
            a.get("branch_name") or "-",
            STATUS_LABELS.get(a.get("status"), a.get("status") or "-"),
            a.get("phone") or "-",
            a.get("city") or "-",
            a.get("district") or "-",
            a.get("birth_date") or "-",
            a.get("education") or "-",
            a.get("exp_years") or "-",
            uniform.get(a.get("uniform_status"), "Noma'lum"),
            a.get("expected_salary") or "-",
            a.get("created_at") or "-",
        ])
    _write_sheet(ws, headers, rows)
    return _finish(wb, "arizalar")


def build_users_xlsx(users):
    wb = Workbook()
    ws = wb.active
    ws.title = "Foydalanuvchilar"
    headers = ["#", "TG ID", "Ism-sharif", "Username", "Telefon", "Rol", "Filial", "Holat", "Sana"]
    rows = []
    for u in users:
        rows.append([
            u.get("id"),
            u.get("tg_id"),
            u.get("full_name") or "-",
            ("@" + u["username"]) if u.get("username") else "-",
            u.get("phone") or "-",
            u.get("role") or "-",
            u.get("branch_name") or "-",
            "Bloklangan" if u.get("blocked") else "Faol",
            u.get("created_at") or "-",
        ])
    _write_sheet(ws, headers, rows)
    return _finish(wb, "foydalanuvchilar")


def build_advance_xlsx(rows, period, pay_date=None):
    """Avans oluvchilar ro'yxati — chiroyli, tushunarli dizayn bilan.

    rows — list_advances() natijasi (ism, karta, filial, lavozim, telefon).
    period — 'YYYY-MM'. pay_date — to'lov sanasi (masalan '15.07.2026').
    """
    thin = Side(style="thin", color="BBBBBB")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    title_fill = PatternFill("solid", fgColor="1B5E20")
    title_font = Font(bold=True, color="FFFFFF", size=16)
    sub_font = Font(bold=True, color="1B5E20", size=11)
    alt_fill = PatternFill("solid", fgColor="E8F5E9")
    center = Alignment(horizontal="center", vertical="center")
    left = Alignment(horizontal="left", vertical="center")

    headers = ["№", "Ism-familiya", "Karta raqami", "Filial", "Lavozim", "Telefon"]
    ncols = len(headers)
    last_col = get_column_letter(ncols)

    wb = Workbook()
    ws = wb.active
    ws.title = "Avans"
    ws.sheet_view.showGridLines = False

    # 1-qator: sarlavha banneri
    ws.merge_cells(f"A1:{last_col}1")
    c = ws["A1"]
    c.value = "GULNORA FARM — AVANS OLUVCHILAR RO'YXATI"
    c.fill = title_fill
    c.font = title_font
    c.alignment = center
    ws.row_dimensions[1].height = 30

    # 2-qator: davr / to'lov sanasi / jami
    ws.merge_cells(f"A2:{last_col}2")
    info = f"Davr: {period}"
    if pay_date:
        info += f"    |    To'lov sanasi: {pay_date}"
    info += f"    |    Jami: {len(rows)} nafar"
    c2 = ws["A2"]
    c2.value = info
    c2.font = sub_font
    c2.alignment = center
    ws.row_dimensions[2].height = 20

    # 3-qator: ustun sarlavhalari
    ws.append([])  # 3-qatorga o'tkazish uchun bo'sh — keyin qo'lda yozamiz
    header_row = 3
    for i, h in enumerate(headers, start=1):
        cell = ws.cell(row=header_row, column=i, value=h)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = center
        cell.border = border
    ws.row_dimensions[header_row].height = 22

    # Ma'lumot qatorlari
    for idx, r in enumerate(rows, start=1):
        row_i = header_row + idx
        values = [
            idx,
            r.get("full_name") or "-",
            r.get("card_number") or "-",
            r.get("branch_name") or "-",
            r.get("position") or "-",
            r.get("phone") or "-",
        ]
        for col_i, val in enumerate(values, start=1):
            cell = ws.cell(row=row_i, column=col_i, value=val)
            cell.border = border
            cell.alignment = center if col_i in (1,) else left
            if idx % 2 == 0:
                cell.fill = alt_fill
        # karta raqami matn sifatida (raqam formatlanmasin)
        ws.cell(row=row_i, column=3).number_format = "@"

    # Ustun kengliklari
    widths = [6, 28, 24, 26, 20, 18]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.freeze_panes = f"A{header_row + 1}"
    return _finish(wb, f"avans_{period}")


def _styled_header_row(ws, row_i, headers, border, center):
    for i, h in enumerate(headers, start=1):
        cell = ws.cell(row=row_i, column=i, value=h)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = center
        cell.border = border
    ws.row_dimensions[row_i].height = 22


def build_dayoff_xlsx(branches_data, date_display):
    """Kunlik dam olish hisoboti — filial bo'lim-bo'lim, chiroyli dizayn.

    branches_data — [{branch_name, items:[{full_name, position, day_status}]}].
    """
    thin = Side(style="thin", color="BBBBBB")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    title_fill = PatternFill("solid", fgColor="1B5E20")
    title_font = Font(bold=True, color="FFFFFF", size=16)
    sub_font = Font(bold=True, color="1B5E20", size=11)
    branch_fill = PatternFill("solid", fgColor="C8E6C9")
    branch_font = Font(bold=True, color="1B5E20", size=12)
    off_fill = PatternFill("solid", fgColor="FFF3E0")
    center = Alignment(horizontal="center", vertical="center")
    left = Alignment(horizontal="left", vertical="center")

    headers = ["№", "Ism-familiya", "Lavozim", "Holat"]
    ncols = len(headers)
    last_col = get_column_letter(ncols)

    wb = Workbook()
    ws = wb.active
    ws.title = "Dam olish"
    ws.sheet_view.showGridLines = False

    ws.merge_cells(f"A1:{last_col}1")
    c = ws["A1"]
    c.value = "GULNORA FARM — KUNLIK DAM OLISH HISOBOTI"
    c.fill = title_fill
    c.font = title_font
    c.alignment = center
    ws.row_dimensions[1].height = 30

    total_off = sum(
        sum(1 for it in b["items"] if it.get("day_status") == "off")
        for b in branches_data
    )
    ws.merge_cells(f"A2:{last_col}2")
    c2 = ws["A2"]
    c2.value = (f"Sana: {date_display}    |    Filiallar: {len(branches_data)}"
                f"    |    Jami dam oluvchi: {total_off} nafar")
    c2.font = sub_font
    c2.alignment = center
    ws.row_dimensions[2].height = 20

    row_i = 4
    for b in branches_data:
        off_items = [it for it in b["items"] if it.get("day_status") == "off"]
        # Filial sarlavhasi
        ws.merge_cells(start_row=row_i, start_column=1, end_row=row_i, end_column=ncols)
        bc = ws.cell(row=row_i, column=1,
                     value=f"🏢 {b['branch_name']}  —  dam oluvchi: {len(off_items)} nafar")
        bc.fill = branch_fill
        bc.font = branch_font
        bc.alignment = left
        for col_i in range(1, ncols + 1):
            ws.cell(row=row_i, column=col_i).border = border
        ws.row_dimensions[row_i].height = 20
        row_i += 1
        # Ustun sarlavhalari
        _styled_header_row(ws, row_i, headers, border, center)
        row_i += 1
        if not off_items:
            ws.merge_cells(start_row=row_i, start_column=1, end_row=row_i, end_column=ncols)
            ec = ws.cell(row=row_i, column=1, value="— Bu filialda ertaga dam oluvchi yo'q —")
            ec.alignment = center
            for col_i in range(1, ncols + 1):
                ws.cell(row=row_i, column=col_i).border = border
            row_i += 2
            continue
        for idx, it in enumerate(off_items, start=1):
            values = [idx, it.get("full_name") or "-", it.get("position") or "-", "🛌 Dam oladi"]
            for col_i, val in enumerate(values, start=1):
                cell = ws.cell(row=row_i, column=col_i, value=val)
                cell.border = border
                cell.alignment = center if col_i in (1, 4) else left
                cell.fill = off_fill
            row_i += 1
        row_i += 1  # bo'lim orasida bo'sh qator

    widths = [6, 30, 24, 16]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    return _finish(wb, "dam_olish")


def build_daily_attendance_xlsx(branches_data, date_display):
    """Direktor uchun kunlik davomat hisoboti — filial bo'lim-bo'lim.

    branches_data — [{branch_name, total, present, absent,
                      detail:[{full_name, came, out, late, early}]}].
    """
    thin = Side(style="thin", color="BBBBBB")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    title_fill = PatternFill("solid", fgColor="0D47A1")
    title_font = Font(bold=True, color="FFFFFF", size=16)
    sub_font = Font(bold=True, color="0D47A1", size=11)
    branch_fill = PatternFill("solid", fgColor="BBDEFB")
    branch_font = Font(bold=True, color="0D47A1", size=12)
    alt_fill = PatternFill("solid", fgColor="E3F2FD")
    center = Alignment(horizontal="center", vertical="center")
    left = Alignment(horizontal="left", vertical="center")

    headers = ["№", "Ism-familiya", "Keldi", "Ketdi", "Izoh"]
    ncols = len(headers)
    last_col = get_column_letter(ncols)

    wb = Workbook()

    # 1-varaq: Umumiy jadval (filiallar kesimi)
    ws0 = wb.active
    ws0.title = "Umumiy"
    ws0.sheet_view.showGridLines = False
    ws0.merge_cells("A1:D1")
    c = ws0["A1"]
    c.value = "GULNORA FARM — KUNLIK DAVOMAT HISOBOTI"
    c.fill = title_fill
    c.font = title_font
    c.alignment = center
    ws0.row_dimensions[1].height = 30
    ws0.merge_cells("A2:D2")
    c2 = ws0["A2"]
    c2.value = f"Sana: {date_display}"
    c2.font = sub_font
    c2.alignment = center
    _styled_header_row(ws0, 4, ["Filial", "Jami xodim", "Kelgan", "Kelmagan"], border, center)
    r = 5
    tot_e = tot_p = tot_a = 0
    for b in branches_data:
        tot_e += b["total"]; tot_p += b["present"]; tot_a += b["absent"]
        vals = [b["branch_name"], b["total"], b["present"], b["absent"]]
        for col_i, val in enumerate(vals, start=1):
            cell = ws0.cell(row=r, column=col_i, value=val)
            cell.border = border
            cell.alignment = left if col_i == 1 else center
            if r % 2 == 0:
                cell.fill = alt_fill
        r += 1
    for col_i, val in enumerate(["JAMI", tot_e, tot_p, tot_a], start=1):
        cell = ws0.cell(row=r, column=col_i, value=val)
        cell.font = Font(bold=True)
        cell.border = border
        cell.alignment = left if col_i == 1 else center
    for i, w in enumerate([28, 14, 12, 12], start=1):
        ws0.column_dimensions[get_column_letter(i)].width = w

    # 2-varaq: Filial bo'lim-bo'lim tafsilot
    ws = wb.create_sheet("Tafsilot")
    ws.sheet_view.showGridLines = False
    row_i = 1
    for b in branches_data:
        ws.merge_cells(start_row=row_i, start_column=1, end_row=row_i, end_column=ncols)
        bc = ws.cell(
            row=row_i, column=1,
            value=(f"🏢 {b['branch_name']}  —  Jami: {b['total']} | "
                   f"Kelgan: {b['present']} | Kelmagan: {b['absent']}"),
        )
        bc.fill = branch_fill
        bc.font = branch_font
        bc.alignment = left
        for col_i in range(1, ncols + 1):
            ws.cell(row=row_i, column=col_i).border = border
        ws.row_dimensions[row_i].height = 20
        row_i += 1
        _styled_header_row(ws, row_i, headers, border, center)
        row_i += 1
        detail = b.get("detail") or []
        if not detail:
            ws.merge_cells(start_row=row_i, start_column=1, end_row=row_i, end_column=ncols)
            ec = ws.cell(row=row_i, column=1, value="— Bugun hech kim kelmagan —")
            ec.alignment = center
            for col_i in range(1, ncols + 1):
                ws.cell(row=row_i, column=col_i).border = border
            row_i += 2
            continue
        for idx, d in enumerate(detail, start=1):
            note = []
            if d.get("late"):
                note.append("kech keldi")
            if d.get("early"):
                note.append("erta ketdi")
            values = [
                idx, d.get("full_name") or "-", d.get("came") or "-",
                d.get("out") or "hali ishda", ", ".join(note) or "-",
            ]
            for col_i, val in enumerate(values, start=1):
                cell = ws.cell(row=row_i, column=col_i, value=val)
                cell.border = border
                cell.alignment = center if col_i in (1, 3, 4) else left
                if idx % 2 == 0:
                    cell.fill = alt_fill
            row_i += 1
        row_i += 1
    for i, w in enumerate([6, 30, 12, 14, 22], start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    return _finish(wb, "kunlik_davomat")


def build_report_xlsx(stats, branches, vacancies):
    """Umumiy hisobot: statistika + filiallar + lavozimlar kesimi."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Umumiy"
    _write_sheet(ws, ["Ko'rsatkich", "Qiymat"], [
        ["Bugungi arizalar", stats.get("today", 0)],
        ["Haftalik arizalar", stats.get("week", 0)],
        ["Oylik arizalar", stats.get("month", 0)],
        ["Yangi", stats.get("new", 0)],
        ["Suhbatga chaqirilgan", stats.get("interview", 0)],
        ["Qabul qilingan", stats.get("accepted", 0)],
        ["Rad etilgan", stats.get("rejected", 0)],
        ["Jami arizalar", stats.get("total", 0)],
    ])

    ws_b = wb.create_sheet("Filiallar")
    _write_sheet(ws_b, ["Filial", "Arizalar soni"],
                 [[b.get("name") or "Nomsiz", b.get("cnt", 0)] for b in branches])

    ws_v = wb.create_sheet("Lavozimlar")
    _write_sheet(ws_v, ["Lavozim", "Arizalar soni"],
                 [[v.get("name") or "Nomsiz", v.get("cnt", 0)] for v in vacancies])

    return _finish(wb, "hisobot")
